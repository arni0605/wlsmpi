import json
import io
from collections import OrderedDict
from itertools import islice
from openpyxl import load_workbook
from openpyxl import load_workbook
# load the workbook
wb = load_workbook('wl_mpi_template_v2.xlsx')
file1 = open("input_v3.json","w")
file1.write("{\n")
# Print the sheet names
#print(wb.sheetnames)
# Get a sheet by name
for sheets in wb.sheetnames:
  #print (sheets)
  sheet=wb[sheets]
  i=0
  for row in sheet.values:
     # print(row)
     if(i>=1 and row[6] is not None):
       #print("\"",row[0],"\":\"",row[6],"\",")
       file1.write("\""+str(row[0])+"\":\""+str(row[6])+"\""+",\n")
      # if(i<sheet.max_row-1):
       #   strValue=",\n"
       #else:
       #   strValue="\n"
      # file1.write(strValue)
     i=i+1

file1.write("}")
file1.close()
